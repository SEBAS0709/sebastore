
import csv
import io
from datetime import datetime

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

ENCABEZADOS = [
    "Juego", "Tienda", "Precio normal", "Precio oferta",
    "% Descuento", "Calificación", "Actualizado",
]


def _filas_desde_ofertas(ofertas):
    filas = []
    for o in ofertas:
        filas.append([
            o.juego.titulo if o.juego else "-",
            o.tienda.nombre if o.tienda else "-",
            f"{o.precio_normal:.2f}",
            f"{o.precio_oferta:.2f}",
            f"{o.porcentaje_descuento:.0f}%",
            f"{o.calificacion_oferta:.1f}" if o.calificacion_oferta else "-",
            o.actualizada_en.strftime("%Y-%m-%d %H:%M") if o.actualizada_en else "-",
        ])
    return filas


def exportar_csv(ofertas) -> io.BytesIO:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(ENCABEZADOS)
    writer.writerows(_filas_desde_ofertas(ofertas))

    bytes_buffer = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
    bytes_buffer.seek(0)
    return bytes_buffer


def exportar_excel(ofertas) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ofertas"
    ws.append(ENCABEZADOS)
    for col in range(1, len(ENCABEZADOS) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    for fila in _filas_desde_ofertas(ofertas):
        ws.append(fila)

    for columna in ws.columns:
        longitud = max(len(str(c.value)) for c in columna if c.value is not None)
        ws.column_dimensions[columna[0].column_letter].width = longitud + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_pdf(ofertas) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph("Reporte de ofertas de videojuegos", estilos["Title"]),
        Paragraph(
            f"Generado el {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            estilos["Normal"],
        ),
        Spacer(1, 12),
    ]

    datos_tabla = [ENCABEZADOS] + _filas_desde_ofertas(ofertas)
    tabla = Table(datos_tabla, repeatRows=1)
    tabla.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#181B21")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F1EA")]),
        ])
    )
    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)
    return buffer
